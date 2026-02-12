"""Report Formatter — genera file PDF, XLSX, CSV dai dati strutturati."""
import csv
import io
import os
import uuid
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from app.core.config import settings


UPLOADS_DIR = Path(getattr(settings, "UPLOADS_DIR", "uploads"))


def get_report_dir(company_id: uuid.UUID) -> Path:
    now = datetime.utcnow()
    path = UPLOADS_DIR / "reports" / str(company_id) / str(now.year) / f"{now.month:02d}"
    path.mkdir(parents=True, exist_ok=True)
    return path


def generate_filename(tipo: str, formato: str, company_short: str = "report") -> str:
    data = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    return f"{tipo}_{data}_{company_short}.{formato}"


def format_report(
    data: dict,
    formato: str,
    company_id: uuid.UUID,
    company_name: str = "Azienda",
) -> tuple[str, int]:
    """Format report data to file. Returns (file_path, file_size)."""
    company_short = company_name.split()[0][:20] if company_name else "report"
    filename = generate_filename(data["metadata"].get("titolo", "report").replace(" ", "_").lower(), formato, company_short)
    output_dir = get_report_dir(company_id)
    file_path = str(output_dir / filename)

    if formato == "pdf":
        _format_pdf(data, file_path, company_name)
    elif formato == "xlsx":
        _format_xlsx(data, file_path, company_name)
    elif formato == "csv":
        _format_csv(data, file_path)
    else:
        raise ValueError(f"Formato non supportato: {formato}")

    file_size = os.path.getsize(file_path)
    return file_path, file_size


def _fmt_value(val, fmt: str | None = None) -> str:
    """Format a value for display."""
    if val is None:
        return ""
    if isinstance(val, Decimal):
        if fmt == "currency":
            s = f"{val:,.2f}"
            return s.replace(",", "X").replace(".", ",").replace("X", ".")
        return str(val)
    if isinstance(val, date):
        return val.strftime("%d/%m/%Y")
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y %H:%M")
    if fmt == "percentage":
        return f"{val}%"
    return str(val)


def _format_pdf(data: dict, file_path: str, company_name: str) -> None:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )

    metadata = data.get("metadata", {})
    headers = data.get("headers", [])
    rows = data.get("rows", [])
    totals = data.get("totals", {})
    layout = data.get("layout", {})

    page_size = landscape(A4) if layout.get("orientamento") == "landscape" else A4

    doc = SimpleDocTemplate(
        file_path, pagesize=page_size,
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=20 * mm, bottomMargin=20 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle", parent=styles["Title"],
        fontSize=16, spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "ReportSubtitle", parent=styles["Normal"],
        fontSize=10, textColor=colors.grey, spaceAfter=12,
    )

    elements = []

    # Header
    elements.append(Paragraph(company_name, styles["Normal"]))
    elements.append(Paragraph(metadata.get("titolo", "Report"), title_style))
    if metadata.get("periodo"):
        elements.append(Paragraph(f"Periodo: {metadata['periodo']}", subtitle_style))
    elements.append(Spacer(1, 6 * mm))

    # Table
    if headers and rows:
        header_labels = [h["label"] for h in headers]
        header_names = [h["name"] for h in headers]
        header_fmts = {h["name"]: h.get("format") for h in headers}

        table_data = [header_labels]
        for row in rows:
            table_data.append([
                _fmt_value(row.get(name), header_fmts.get(name))
                for name in header_names
            ])

        # Totals row
        if totals:
            table_data.append([
                _fmt_value(totals.get(name), header_fmts.get(name))
                for name in header_names
            ])

        col_widths = None
        available = page_size[0] - 30 * mm
        if len(headers) > 0:
            col_widths = [available / len(headers)] * len(headers)

        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        style_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]

        # Right-align currency columns
        for i, h in enumerate(headers):
            if h.get("align") == "right":
                style_cmds.append(("ALIGN", (i, 0), (i, -1), "RIGHT"))
            elif h.get("align") == "center":
                style_cmds.append(("ALIGN", (i, 0), (i, -1), "CENTER"))

        # Bold totals row
        if totals:
            last_row = len(table_data) - 1
            style_cmds.append(("FONTNAME", (0, last_row), (-1, last_row), "Helvetica-Bold"))
            style_cmds.append(("BACKGROUND", (0, last_row), (-1, last_row), colors.HexColor("#e8e8e8")))

        table.setStyle(TableStyle(style_cmds))
        elements.append(table)

    # Footer info
    elements.append(Spacer(1, 10 * mm))
    footer_text = f"Generato il {datetime.utcnow().strftime('%d/%m/%Y alle %H:%M')}"
    if metadata.get("numero_movimenti"):
        footer_text += f" — {metadata['numero_movimenti']} movimenti"
    elif metadata.get("numero_conti"):
        footer_text += f" — {metadata['numero_conti']} conti"
    elements.append(Paragraph(footer_text, styles["Normal"]))

    doc.build(elements)


def _format_xlsx(data: dict, file_path: str, company_name: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
    from openpyxl.utils import get_column_letter

    metadata = data.get("metadata", {})
    headers = data.get("headers", [])
    rows = data.get("rows", [])
    totals = data.get("totals", {})

    wb = Workbook()
    ws = wb.active
    ws.title = metadata.get("titolo", "Report")[:31]

    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    total_font = Font(name="Calibri", bold=True, size=10)
    total_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")

    # Title rows
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
    ws.cell(row=1, column=1, value=company_name).font = Font(name="Calibri", bold=True, size=12)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(headers), 1))
    ws.cell(row=2, column=1, value=metadata.get("titolo", "Report")).font = Font(name="Calibri", bold=True, size=14)
    if metadata.get("periodo"):
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max(len(headers), 1))
        ws.cell(row=3, column=1, value=f"Periodo: {metadata['periodo']}").font = Font(name="Calibri", size=10, color="666666")

    start_row = 5
    header_names = [h["name"] for h in headers]
    header_fmts = {h["name"]: h.get("format") for h in headers}

    # Header row
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=h["label"])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Data rows
    for row_idx, row_data in enumerate(rows, start_row + 1):
        for col_idx, name in enumerate(header_names, 1):
            val = row_data.get(name)
            cell = ws.cell(row=row_idx, column=col_idx)

            if isinstance(val, (Decimal, float, int)) and not isinstance(val, bool):
                cell.value = float(val) if isinstance(val, Decimal) else val
                fmt = header_fmts.get(name)
                if fmt == "currency":
                    cell.number_format = '#.##0,00'
                elif fmt == "percentage":
                    cell.number_format = '0.0"%"'
            elif isinstance(val, date):
                cell.value = val
                cell.number_format = "DD/MM/YYYY"
            else:
                cell.value = str(val) if val is not None else ""

            cell.border = border
            align = next((h.get("align", "left") for h in headers if h["name"] == name), "left")
            cell.alignment = Alignment(horizontal=align)

    # Totals row
    if totals:
        total_row = start_row + len(rows) + 1
        for col_idx, name in enumerate(header_names, 1):
            val = totals.get(name)
            cell = ws.cell(row=total_row, column=col_idx)
            if isinstance(val, (Decimal, float, int)) and not isinstance(val, bool):
                cell.value = float(val) if isinstance(val, Decimal) else val
                fmt = header_fmts.get(name)
                if fmt == "currency":
                    cell.number_format = '#.##0,00'
            else:
                cell.value = str(val) if val is not None else ""
            cell.font = total_font
            cell.fill = total_fill
            cell.border = border

    # Auto-width
    for col_idx in range(1, len(headers) + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # Freeze pane
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    # Auto-filter
    if headers:
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A{start_row}:{last_col}{start_row + len(rows)}"

    # Parametri sheet
    ws2 = wb.create_sheet("Parametri")
    ws2.cell(row=1, column=1, value="Parametro").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Valore").font = Font(bold=True)
    for i, (k, v) in enumerate(metadata.items(), 2):
        ws2.cell(row=i, column=1, value=k)
        ws2.cell(row=i, column=2, value=str(v))
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 40

    wb.save(file_path)


def _format_csv(data: dict, file_path: str) -> None:
    headers = data.get("headers", [])
    rows = data.get("rows", [])
    totals = data.get("totals", {})
    header_names = [h["name"] for h in headers]
    header_fmts = {h["name"]: h.get("format") for h in headers}

    with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";", quoting=csv.QUOTE_MINIMAL)

        # Header
        writer.writerow([h["label"] for h in headers])

        # Data rows
        for row_data in rows:
            csv_row = []
            for name in header_names:
                val = row_data.get(name)
                csv_row.append(_fmt_value(val, header_fmts.get(name)))
            writer.writerow(csv_row)

        # Totals
        if totals:
            csv_row = []
            for name in header_names:
                val = totals.get(name)
                csv_row.append(_fmt_value(val, header_fmts.get(name)))
            writer.writerow(csv_row)
